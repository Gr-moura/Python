from datetime import datetime, timedelta
import yfinance as yf
import openpyxl

def ReiniciarPlanilha():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    workbook.save("Planilha.xlsx")

# 1. Função para obter dados históricos de ações com yfinance
def obter_dados_acao(ticker):
    # Calcular a data de início e fim
    fim = datetime.now()
    inicio = fim - timedelta(days=20*365)
    
    # Baixar dados históricos da ação (ex: Apple 'AAPL')
    acao = yf.Ticker(ticker)
    dados = acao.history(start=inicio.strftime('%Y-%m-%d'), end=fim.strftime('%Y-%m-%d'))
    return dados

# 2. Função para salvar os dados em uma planilha Excel usando openpyxl
def salvar_dados_em_excel(dados, nome_arquivo, acao = 'Acao'):
    if dados.empty:
        print(f"Nenhum dado encontrado para salvar em {nome_arquivo}")
        return

    try:
        # Tentar abrir o arquivo Excel existente
        workbook = openpyxl.load_workbook(nome_arquivo)
        sheet = workbook.active
    except FileNotFoundError:
        # Se o arquivo não existir, criar um novo
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        print(f"Arquivo {nome_arquivo} não encontrado.")
        return

    # Encontrar a próxima coluna vazia
    col_offset = sheet.max_column
    
    if (sheet.max_column != 1):
        col_offset = sheet.max_column + 1

    # Escrever cabeçalhos na nova coluna
    headers = ["Data", "Abertura", "Alta", "Baixa", "Fechamento", "Volume", "Splits", "Compras", "Lucro"]
    for i, header in enumerate(headers):
        sheet.cell(row=1, column=col_offset + i, value=header)

    # Escrever os dados históricos na planilha
    for row_idx, (index, row) in enumerate(dados.iterrows(), start=2):
        data = index.strftime('%Y-%m-%d')
        valores = [data, row['Open'], row['High'], row['Low'], row['Close'], row['Volume'], row['Stock Splits']]
        for col_idx, valor in enumerate(valores):
            sheet.cell(row=row_idx, column=col_offset + col_idx, value=valor)


    # Salvar o arquivo Excel
    workbook.save(nome_arquivo)
    print(f"{acao} salva em {nome_arquivo}")
