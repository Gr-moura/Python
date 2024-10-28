from openpyxl import load_workbook
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import Compra

#Essa função retorna uma lista com os lucros obtidos em cada ação durante o backtest

#Também assumi que qualquer ordem de compra dada pela função compra significa comprar a ação e vender
#ela no dia seguinte. Isto porque é sempre melhor vender todas as ações antes de rodar o programa, 
#para que ele calcule novamente qual é o melhor uso do dinheiro investido. Se necessário, o programa pode
#simplesmente recomprar todas as ações após a venda. Qualquer coisa me pergunta se nao entender
def runBacktest(start, end, numberOfStocks):
    file = 'Planilha.xlsx'
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    startRow = 0
    endRow = 0
    for row in ws.rows:
        if(row[0].value == start):
            startRow = row[0].row

        if(row[0].value == end):
            endRow = row[0].row
            break

    for row in range(startRow, endRow):
        compras = Compra.compra(row)
        for i in range(len(compras)):
            ws.cell(row=row, column=(9 * (i + 1)) - 1, value=compras[i])

    totalProfit = [0] * numberOfStocks
    for row in range(startRow, endRow):
        for i in range(numberOfStocks):
            profit = ws.cell(row=row + 1, column=(9 * (i + 1)) - 4).value - ws.cell(row=row, column=(9 * (i + 1)) - 4).value
            profit *= ws.cell(row=row, column=(9 * (i + 1)) - 1).value

            totalProfit[i] += profit
            ws.cell(row=row, column=(9 * (i + 1)), value=profit,)

            if(profit > 0):
                ws.cell(row, (9 * (i + 1))).fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type="solid")
            elif(profit < 0):
                ws.cell(row, (9 * (i + 1))).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
            else:
                ws.cell(row, (9 * (i + 1))).fill = PatternFill(start_color='A8A8A8', end_color='A8A8A8', fill_type="solid")

    wb.save(file)
    return totalProfit
