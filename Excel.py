from datetime import datetime, timedelta
import yfinance as yf
import openpyxl

Lista = ["BBS3.sa", "PETR4.sa", "VALE3.sa", "ITUB4.sa", "BBDC4.sa", "ABEV3.sa", "WEGE3.sa", "VVAR3.sa", "MGLU3.sa", "BBAS3.sa", "GNDI3.sa", "LREN3.sa", "RENT3.sa", "JBSS3.sa", "NTCO3.sa", "HAPV3.sa", "CSNA3.sa", "IRBR3.sa", "BRFS3.sa", "CIEL3.sa", "ELET3.sa", "ELET6.sa", "CMIG4.sa", "SBSP3.sa", "BRKM5.sa", "USIM5.sa", "GGBR4.sa", "CSAN3.sa", "MRFG3.sa", "BTOW3.sa", "BRML3.sa", "CYRE3.sa", "MULT3.sa", "HYPE3.sa", "LAME4.sa", "QUAL3.sa", "BRAP4.sa", "CCRO3.sa", "ENBR3.sa", "EGIE3.sa", "CPFE3.sa", "TAEE11.sa", "SAPR4.sa", "CVCB3.sa", "AZUL4.sa", "GOLL4.sa", "B3SA3.sa", "IRBR3.sa", "BBSE3.sa", "SULA11.sa", "BBDC3.sa", "BBDC4.sa", "ITUB3.sa", "ITUB4.sa", "SANB11.sa", "BBAS3.sa", "BPAC11.sa", "BIDI11.sa", "BIDI4.sa", "BIDI3.sa", "BRSR6.sa", "ITSA3.sa", "ITSA4.sa", "ITUB3.sa", "ITUB4.sa", "BBDC3.sa", "BBDC4.sa", "SANB11.sa", "BBAS3.sa", "BPAC11.sa", "BIDI11.sa", "BIDI4.sa", "BIDI3.sa", "BRSR6.sa", "ITSA3.sa", "ITSA4.sa", "ITUB3.sa", "ITUB4.sa", "BBDC3.sa", "BBDC4.sa", "SANB11.sa", "BBAS3.sa", "BPAC11.sa",]

def ReiniciarPlanilha():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    workbook.save("Planilha.xlsx")

# 1. Função para obter dados históricos de ações com yfinance
def obter_dados_acao(ticker):
    # Calcular a data de início e fim
    fim = datetime.now()
    inicio = fim - timedelta(days=20*365)
    
    # Baixar dados históricos da ação (ex: Apple 'AAPL')
    acao = yf.Ticker(ticker)
    dados = acao.history(start=inicio.strftime('%Y-%m-%d'), end=fim.strftime('%Y-%m-%d'))
    return dados

# 2. Função para salvar os dados em uma planilha Excel usando openpyxl
def salvar_dados_em_excel(dados, nome_arquivo):
    if dados.empty:
        print(f"Nenhum dado encontrado para salvar em {nome_arquivo}")
        return

    try:
        # Tentar abrir o arquivo Excel existente
        workbook = openpyxl.load_workbook(nome_arquivo)
        sheet = workbook.active
    except FileNotFoundError:
        # Se o arquivo não existir, criar um novo
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        print(f"Arquivo {nome_arquivo} não encontrado.")
        return

    # Encontrar a próxima coluna vazia
    col_offset = sheet.max_column
    
    if (sheet.max_column != 1):
        col_offset = sheet.max_column + 2 

    # Escrever cabeçalhos na nova coluna
    headers = ["Data", "Abertura", "Alta", "Baixa", "Fechamento", "Volume"]
    for i, header in enumerate(headers):
        sheet.cell(row=1, column=col_offset + i, value=header)

    # Escrever os dados históricos na planilha
    for row_idx, (index, row) in enumerate(dados.iterrows(), start=2):
        data = index.strftime('%Y-%m-%d')
        valores = [data, row['Open'], row['High'], row['Low'], row['Close'], row['Volume']]
        for col_idx, valor in enumerate(valores):
            sheet.cell(row=row_idx, column=col_offset + col_idx, value=valor)


    # Salvar o arquivo Excel
    workbook.save(nome_arquivo)
    print(f"Dados salvos com sucesso em {nome_arquivo}")

# Exemplo de uso
ReiniciarPlanilha()
for ticker in Lista:
    dados = obter_dados_acao(ticker)
    salvar_dados_em_excel(dados, "Planilha.xlsx")
